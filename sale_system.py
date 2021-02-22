from openpyxl import load_workbook
import time,random,sys,cv2,json
import pyzbar.pyzbar as pyzbar

tcode = ''
sys.setrecursionlimit(100000) #设置递归深度
now_time = time.strftime("%m-%d %H:%M:%S", time.localtime(time.time()))
token = str(int(time.time())) + f'{str(random.randint(0, 9))}{str(random.randint(0, 9))}'

# -------------------------------文件路径注意修改-------------------------------------
wb = load_workbook(r'C:/Users/Administrator/Desktop/1234.xlsx') #手动创建
# ---------------------------------------------------------------------------------

sheet = wb['Sheet1']
sheet['A1'].value = '商品名称'
sheet['B1'].value = '商品销量(个)'
sheet['C1'].value = '商品单价(元)'
sheet['D1'].value = '下单时间'
sheet['E1'].value = '订单号'
sheet['F1'].value = '宿舍号'
sheet['G1'].value = '修改时间'

def sum():
    total = 0
    for index, rows in enumerate(sheet):
        if index !=0:
            total+=int(rows[1].value)*int(rows[2].value)
    print(f'小小食堂到目前为止，合计收入{total}元。')
def modification(name,num):
    for index,rows in enumerate(sheet):
        if index==1:    #获取行标,第index行.
            rows[0].value = name  #name
            rows[1].value = num  #num
            # rows[2].value = '3'    #单价
            # rows[5].value = '3'    #宿舍号
            # rows[3].value
            rows[6].value = now_time
            print('修改成功!')
            
def addition(data): #data为列表传入
    # token = str(int(time.time())) + f'{str(random.randint(0, 9))}{str(random.randint(0, 9))}'
    # data[4] = token
    for i in data:
        sheet.append(i)
        print('添加成功!')

def delete_all():
    msg = input('警告！您即将删除所有数据,继续请按(y/n):')
    if msg=='y':
        # sheet.max_row   获取最大行数
        sheet.delete_rows(2,sheet.max_row)
        print('所有数据删除成功！')
    else:
        print('您取消了操作！')
def print_data(d):
    for row in sheet.values:
        # print(row[4])
        if row[4]==d:
            print(row)



def loading():
    print('\n开发者:阿狸')
    print('*' * 10 + '欢迎来到收银管理系统' + '*' * 10)
    print('1.下单')
    print('2.删除订单')
    print('3.查找订单')
    print('4.修改订单')
    print('0.退出程序')
    print('*' * 10 + '欢迎来到收银管理系统' + '*' * 10)
    sum()
# ------------------调入条码扫描-----------------------
def camera():
    # ---------------------------<ip摄像头>app的公网地址----------------------------------------------------
    camera = cv2.VideoCapture('rtsp://admin:admin@192.168.5.47:8554/live')#默认0为电脑的摄像头
    while True:
        # 读取当前帧
        ret, img = camera.read()
        decode(img)
        if tcode !='':
            camera.release()
            cv2.destroyAllWindows()
            break
        img = cv2.resize(img, None, fx=0.5, fy=0.5, interpolation=cv2.INTER_CUBIC)
        cv2.imshow("camera", img)
        # 按q键退出程序
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    camera.release()
    cv2.destroyAllWindows()


def decode(image):  # 解码
    global tcode
    barcodes = pyzbar.decode(image)
    for barcode in barcodes:
        # 提取并绘制图像中条形码的边界框
        (x, y, w, h) = barcode.rect
        cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)
        barcodeData = barcode.data.decode("utf-8")
        barcodeType = barcode.type
        # 绘出图像上条形码
        if barcodeData:
            tcode+=str(barcodeData)

def find():
    # --------------------注意文件路径修改--------------------------------------
    with open('商品条码数据.json', 'r', encoding='utf-8')as file:
    # --------------------注意文件路径修改--------------------------------------
        old = json.load(file)
        for key,value in old.items():
            if tcode in value:
                print('识别成功！')
                return key
            else:
                print('本地数据库中查找不到此商品，请添加商品信息：')
                pass







def main():
    loading()
    msg = input('请输入序号:')
    if msg=='1':
        camera()
        name = find()
        nandp = input('请输入数量和宿舍号,用逗号或者空格隔开:')
        num = nandp[:-4]
        place = nandp[-3:]
        print(num,place)
        # ------------------------data为测试数据写入,可以删除--------------------------------
        #   [商品名称,销量,单价,下单时间,订单号,宿舍号,修改时间]
        #   sales,place 为用户输入参数
        data = [
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None],
            [name, num, 1, now_time, token, place, None]
        ]
        # ------------------------data为测试数据写入,可以删除--------------------------------
        addition(data)
        main()
    elif msg=='2':
        delete_all()
        main()
    elif msg=='3':
        d = input('请输入订单号:')
        print_data(d)
        main()
    elif msg =='4':
        name = input('暂仅支持修改商品名和数量\n请输入商品名称:')
        num = input('请输入数量:')
        modification(name,num)
        main()
    elif msg=='0':
        pass

    # -----------------------注意同上路径保持一致------------------------------------
    wb.save('C:/Users/Administrator/Desktop/1234.xlsx')
    # -----------------------注意同上路径保持一致------------------------------------

if __name__=='__main__':
    print("程序启动中,请稍等:")
    for i in range(26):
        a = "▋" * i
        b = " " * (25 - i)
        c = (i / 25) * 25 * 4
        print("\r{}{}{:.2f}%".format(a, b, c), end="")
        time.sleep(0.1)
    main()

